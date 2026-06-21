import io
from datetime import datetime, timedelta

import openpyxl
from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException, Query, File, UploadFile
from fastapi.responses import StreamingResponse

from app.database import db
from app.models.job import JobCreate, JobResponse, JobStats, JobUpdate
from app.services.auth_service import get_current_user

router = APIRouter(prefix='/api/jobs', tags=['jobs'])


def _doc_to_response(doc: dict) -> dict:
    """Convert a MongoDB job document to a response-friendly dict."""
    job = {**doc}
    job['id'] = str(job.pop('_id'))
    job.pop('user_id', None)
    return job


@router.get('/stats', response_model=JobStats)
async def get_job_stats(current_user: dict = Depends(get_current_user)):
    """Get aggregated job application statistics."""
    user_id = current_user['id']

    pipeline = [
        {'$match': {'user_id': user_id}},
        {'$group': {
            '_id': '$status',
            'count': {'$sum': 1},
        }},
    ]

    status_counts = {'applied': 0, 'interview': 0, 'offer': 0, 'rejected': 0}
    cursor = await db.job_applications.aggregate(pipeline)
    async for doc in cursor:
        if doc['_id'] in status_counts:
            status_counts[doc['_id']] = doc['count']

    total = sum(status_counts.values())
    interview_rate = (status_counts['interview'] / total * 100) if total > 0 else 0.0
    offer_rate = (status_counts['offer'] / total * 100) if total > 0 else 0.0

    return JobStats(
        total=total,
        applied=status_counts['applied'],
        interview=status_counts['interview'],
        offer=status_counts['offer'],
        rejected=status_counts['rejected'],
        interview_rate=round(interview_rate, 1),
        offer_rate=round(offer_rate, 1),
    )


@router.get('/')
async def get_jobs(
    status: str = Query(None),
    job_type: str = Query(None),
    search: str = Query(None),
    sort: str = Query('newest'),
    current_user: dict = Depends(get_current_user),
):
    """Get job applications with optional filters and sorting."""
    user_id = current_user['id']
    query = {'user_id': user_id}

    if status:
        query['status'] = status

    if job_type:
        query['job_type'] = job_type

    if search:
        # Case-insensitive regex search on company and role
        query['$or'] = [
            {'company': {'$regex': search, '$options': 'i'}},
            {'role': {'$regex': search, '$options': 'i'}},
        ]

    # Determine sort order
    if sort == 'oldest':
        sort_key = [('application_date', 1)]
    elif sort == 'company':
        sort_key = [('company', 1)]
    else:  # newest (default)
        sort_key = [('application_date', -1)]

    cursor = db.job_applications.find(query).sort(sort_key)
    jobs = []
    async for doc in cursor:
        jobs.append(_doc_to_response(doc))

    return jobs


def parse_excel_date(value) -> str:
    """Helper to parse dates from Excel which might be datetime objects or strings."""
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=value)).strftime('%Y-%m-%d')
        except:
            pass
    if isinstance(value, str):
        val = value.replace('0205', '2025').strip()
        formats = [
            '%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d', 
            '%Y/%m/%d', '%m-%d-%Y', '%d-%m-%Y', 
            '%Y-%m-%d %H:%M:%S', '%m/%d/%y', '%d/%m/%y'
        ]
        for fmt in formats:
            try:
                dt = datetime.strptime(val, fmt)
                return dt.strftime('%Y-%m-%d')
            except ValueError:
                continue
    return datetime.utcnow().strftime('%Y-%m-%d')


@router.get('/export')
async def export_jobs(current_user: dict = Depends(get_current_user)):
    """Export all user jobs to an Excel file."""
    user_id = current_user['id']
    cursor = db.job_applications.find({'user_id': user_id}).sort('created_at', -1)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Applications"
    
    headers = ['Company', 'Role', 'Date', 'Selection', 'Oncampus/Offcampus', 'Notes']
    ws.append(headers)
    
    async for doc in cursor:
        status_map = {'applied': 'Applied', 'interview': 'Interview', 'offer': 'Selected', 'rejected': 'Not Selected'}
        type_map = {'on-campus': 'OnCampus', 'off-campus': 'OffCampus'}
        
        ws.append([
            doc.get('company', ''),
            doc.get('role', ''),
            doc.get('application_date', ''),
            status_map.get(doc.get('status', 'applied'), 'Applied'),
            type_map.get(doc.get('job_type', 'off-campus'), 'OffCampus'),
            doc.get('notes', '')
        ])
    
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="Pratham_Job_Tracker.xlsx"'}
    )


@router.post('/import')
async def import_jobs(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Import jobs from an uploaded Excel file."""
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
        
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail="Invalid Excel file format")
        
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) <= 1:
        return {"message": "No data found in file", "imported": 0}
        
    data_rows = rows[1:] # Skip header row
    headers = [str(h).strip().lower() if h else '' for h in rows[0]]
    
    col_map = {}
    for idx, h in enumerate(headers):
        if not h: continue
        if 'company' in h: col_map['company'] = idx
        elif 'role' in h or 'title' in h: col_map['role'] = idx
        elif 'date' in h or 'applied' in h: col_map['date'] = idx
        elif 'status' in h or 'selection' in h: col_map['status'] = idx
        elif 'type' in h or 'campus' in h: col_map['type'] = idx
        elif 'location' in h: col_map['location'] = idx
        elif 'url' in h or 'link' in h: col_map['url'] = idx
        elif 'salary' in h: col_map['salary'] = idx
        elif 'note' in h: col_map['notes'] = idx
        elif 'follow' in h or 'next' in h: col_map['follow'] = idx

    imported_count = 0
    now = datetime.utcnow()
    
    for row in data_rows:
        def get_val(key):
            idx = col_map.get(key)
            if idx is not None and len(row) > idx and row[idx] is not None:
                val_str = str(row[idx]).strip()
                return val_str if val_str else None
            return None

        company = get_val('company')
        if not company:
            continue
            
        role = get_val('role') or 'Unknown Role'
        date_val = get_val('date')
        selection_val = (get_val('status') or '').lower()
        type_val = (get_val('type') or '').lower()
        loc_val = (get_val('location') or '').lower()
        notes = get_val('notes')
        url = get_val('url')
        salary = get_val('salary')
        
        parsed_date = parse_excel_date(date_val) if date_val else datetime.utcnow().strftime('%Y-%m-%d')
        
        status = 'applied'
        if 'not selected' in selection_val or 'reject' in selection_val:
            status = 'rejected'
        elif 'selected' in selection_val or 'offer' in selection_val:
            status = 'offer'
        elif 'interview' in selection_val:
            status = 'interview'
            
        job_type = None
        if type_val:
            job_type = 'on-campus' if 'oncampus' in type_val.replace('-', '') else 'off-campus'
            
        location = None
        if loc_val:
            if 'onsite' in loc_val or 'office' in loc_val: location = 'onsite'
            elif 'hybrid' in loc_val: location = 'hybrid'
            else: location = 'remote'
        elif job_type == 'on-campus':
            location = 'onsite'
        
        # Deduplication Check (Option B)
        existing = await db.job_applications.find_one({
            'user_id': current_user['id'],
            'company': {'$regex': f'^{company}$', '$options': 'i'},
            'role': {'$regex': f'^{role}$', '$options': 'i'}
        })
        
        if existing:
            continue
            
        job_doc = {
            'company': company,
            'role': role,
            'application_date': parsed_date,
            'status': status,
            'job_type': job_type,
            'location': location,
            'job_url': url,
            'salary_range': salary,
            'notes': notes,
            'follow_up_date': None,
            'user_id': current_user['id'],
            'follow_up_email_sent': False,
            'created_at': now,
            'updated_at': now,
        }
        
        await db.job_applications.insert_one(job_doc)
        imported_count += 1
        
    return {"message": f"Successfully imported {imported_count} applications", "imported": imported_count}


@router.post('/', response_model=JobResponse, status_code=201)
async def create_job(
    job_data: JobCreate,
    current_user: dict = Depends(get_current_user),
):
    """Create a new job application."""
    now = datetime.utcnow()
    job_doc = {
        **job_data.model_dump(),
        'user_id': current_user['id'],
        'follow_up_email_sent': False,
        'created_at': now,
        'updated_at': now,
    }

    result = await db.job_applications.insert_one(job_doc)
    job_doc['_id'] = result.inserted_id

    return _doc_to_response(job_doc)


@router.put('/{job_id}', response_model=JobResponse)
async def update_job(
    job_id: str,
    job_data: JobUpdate,
    current_user: dict = Depends(get_current_user),
):
    """Update an existing job application."""
    update_fields = {k: v for k, v in job_data.model_dump().items() if v is not None}

    if not update_fields:
        raise HTTPException(status_code=400, detail='No fields to update')

    update_fields['updated_at'] = datetime.utcnow()

    # If follow_up_date changed, reset follow_up_email_sent
    if 'follow_up_date' in update_fields:
        update_fields['follow_up_email_sent'] = False

    result = await db.job_applications.find_one_and_update(
        {'_id': ObjectId(job_id), 'user_id': current_user['id']},
        {'$set': update_fields},
        return_document=True,
    )

    if not result:
        raise HTTPException(status_code=404, detail='Job application not found')

    return _doc_to_response(result)


@router.delete('/{job_id}')
async def delete_job(
    job_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Delete a job application."""
    result = await db.job_applications.delete_one({
        '_id': ObjectId(job_id),
        'user_id': current_user['id'],
    })

    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail='Job application not found')

    return {'message': 'Job application deleted successfully'}


@router.patch('/{job_id}/status')
async def update_job_status(
    job_id: str,
    status_data: dict,
    current_user: dict = Depends(get_current_user),
):
    """Update only the status of a job application."""
    status = status_data.get('status')
    if status not in ('applied', 'interview', 'offer', 'rejected'):
        raise HTTPException(status_code=400, detail='Invalid status value')

    result = await db.job_applications.find_one_and_update(
        {'_id': ObjectId(job_id), 'user_id': current_user['id']},
        {'$set': {'status': status, 'updated_at': datetime.utcnow()}},
        return_document=True,
    )

    if not result:
        raise HTTPException(status_code=404, detail='Job application not found')

    return _doc_to_response(result)
